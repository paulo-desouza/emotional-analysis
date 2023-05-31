from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import plotly.express as px
from dash import Dash, dcc, html, Input, Output



def count_rows(ws):
    """
    Parameters
    ----------
    ws : worksheet object
    
    Returns
    -------
    number of populated rows in worksheet. 

    """
    count_row = 0

    for row in ws:
        if not all([cell.value is None for cell in row]):
            count_row += 1

    return(count_row)




# Scrape the data from the table at emo.xlsx using openpyxl

wb = load_workbook("emo.xlsx")

ws = wb.active

rc = count_rows(ws)


emo_data = {}
start = 3


for j in range(3, 7):
    for i in range(start, rc+2):
        char = get_column_letter(j)
        
        if i == start:
            emo_data[ws[char+"2"].value] = [ws[char+str(i)].value]
        else: 
            emo_data[ws[char+"2"].value].append(ws[char+str(i)].value)
            


# Plotly.py 

# Data frame: used to represent tabular data. 
# What is the input for panda? What is the function? 

df = pd.DataFrame(data=emo_data)

# ---------------------- #

app = Dash(__name__)

app.layout = html.Div([
    html.H4('Emotional Analysis Graph - Habits'),
    dcc.Graph(id="graph"),
    html.P("Filter by Health:"),
    dcc.RangeSlider(
        id='range-slider',
        min=0, max=10, step=0.5,
        marks={0: '0', 10: "10"},
        value=[0, 10]
    ),
])

@app.callback(
    Output("graph", "figure"), 
    Input("range-slider", "value"))
def update_bar_chart(slider_range):
    low, high = slider_range
    mask = (df.HEALTH > low) & (df.HEALTH < high)

    fig = px.scatter_3d(df[mask], 
        x='ACTIVE', y='HEALTH', z='FUN', 
        color = "HEALTH", hover_data=['HABIT'])
    return fig


app.run_server(debug=True)



# Great Progress! 

# To do: Run this off the web browser;

#        No matter what the filter is at, the ranges should always be at 0 - 10

#        All habits have to show up in the chart. 











